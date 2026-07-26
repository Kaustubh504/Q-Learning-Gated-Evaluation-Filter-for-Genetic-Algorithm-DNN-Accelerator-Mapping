"""
genome_to_excel.py
===================
Reads genome_all.csv (from a single-layer GAMMA run),
counts frequency of each unique genome structure,
and exports to Excel with fitness scores.

Usage:
    python genome_to_excel.py --csv genome_all.csv --out genome_results.xlsx
"""

import argparse
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter


def build_summary(df):
    """
    From raw genome_all.csv:
      - Count how many times each unique genome_str appeared  = frequency
      - Get best fitness for each unique genome_str
      - Get sp_dim (parallelization dimension)
    Returns sorted DataFrame best → worst fitness.
    """
    df["fitness1"] = pd.to_numeric(df["fitness1"], errors="coerce")
    df["valid"]    = pd.to_numeric(df["valid"],    errors="coerce").fillna(0).astype(int)

    # frequency = number of rows with that genome_str
    freq     = df.groupby("genome_str").size().rename("frequency")
    best_fit = df.groupby("genome_str")["fitness1"].max().rename("best_fitness")
    sp_dim   = df.groupby("genome_str")["genome_pattern"].first().rename("sp_dim_pattern")
    valid    = df.groupby("genome_str")["valid"].max().rename("valid")

    summary = pd.concat([freq, best_fit, sp_dim, valid], axis=1).reset_index()
    summary = summary.sort_values("best_fitness", ascending=False).reset_index(drop=True)
    summary.insert(0, "rank", range(1, len(summary) + 1))

    # Verdict: top 20% GOOD, bottom 30% BAD, rest MEDIUM
    n_valid = len(summary[summary["valid"] == 1])
    def get_verdict(row):
        if row["valid"] == 0 or row["best_fitness"] <= -1e17:
            return "INVALID"
        pct = (row["rank"] - 1) / max(n_valid - 1, 1)
        if pct <= 0.20:
            return "GOOD"
        elif pct >= 0.70:
            return "BAD - AVOID"
        else:
            return "MEDIUM"

    summary["verdict"] = summary.apply(get_verdict, axis=1)
    return summary


def parse_genome_str(genome_str):
    """
    Parse 'L1:[sp=X,972] S=3,R=2,K=30,Y=56,X=74,C=2'
    into readable columns: sp_dim, sp_sz, pos1_dim, pos1_tile, ..., pos6_dim, pos6_tile
    """
    result = {}
    try:
        part = genome_str.strip()
        bracket_end = part.index("]")
        sp_part = part[part.index("[")+1: bracket_end]   # "sp=X,972"
        _, sp_rest = sp_part.split("=", 1)
        sp_dim, sp_sz = sp_rest.split(",")
        result["sp_dim"] = sp_dim.strip()
        result["sp_sz"]  = int(sp_sz.strip())

        rest = part[bracket_end+1:].strip()               # "S=3,R=2,K=30,Y=56,X=74,C=2"
        for pos, token in enumerate(rest.split(","), start=1):
            token = token.strip()
            if "=" in token:
                d, sz = token.split("=")
                result[f"pos{pos}_dim"]  = d.strip()
                result[f"pos{pos}_tile"] = int(sz.strip())
    except Exception:
        pass
    return result


def to_excel(summary, out_path):
    wb = Workbook()

    thin  = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_fill = PatternFill("solid", start_color="1F3864")

    verdict_style = {
        "GOOD":        ("E2EFDA", "375623"),
        "MEDIUM":      ("FFF2CC", "7D6608"),
        "BAD - AVOID": ("FCE4D6", "843C0C"),
        "INVALID":     ("D9D9D9", "595959"),
    }

    def write_header(ws, headers, row=1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 28

    def style_cell(cell, verdict, wrap=False):
        bg, fg = verdict_style.get(verdict, ("FFFFFF", "000000"))
        cell.fill   = PatternFill("solid", start_color=bg)
        cell.font   = Font(name="Arial", size=9, color=fg)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=wrap,
                                   horizontal="left" if wrap else "center")

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Genome Summary"

    headers1 = ["Rank", "SP Dim Pattern", "Frequency\n(times appeared)", "Best Fitness",
                "Verdict", "Full Genome Structure (single layer)"]
    write_header(ws1, headers1)

    for ri, row in summary.iterrows():
        er = ri + 2
        vals = [row["rank"], row["sp_dim_pattern"], row["frequency"],
                round(row["best_fitness"], 2) if row["best_fitness"] > -1e17 else "INVALID",
                row["verdict"], row["genome_str"]]
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(row=er, column=ci, value=val)
            style_cell(c, row["verdict"], wrap=(ci == 6))

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 80
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Expanded gene view ───────────────────────────────────────
    ws2 = wb.create_sheet("Gene Detail")
    # Parse each genome_str into individual gene columns
    parsed_rows = []
    for _, row in summary.iterrows():
        # Only use the L1 part (single layer — no ||)
        genome_str = row["genome_str"]
        if "||" in genome_str:
            genome_str = genome_str.split("||")[0].strip()
        genes = parse_genome_str(genome_str)
        r = {
            "rank":          row["rank"],
            "frequency":     row["frequency"],
            "best_fitness":  round(row["best_fitness"], 2) if row["best_fitness"] > -1e17 else "INVALID",
            "verdict":       row["verdict"],
            "sp_dim":        genes.get("sp_dim", ""),
            "sp_sz":         genes.get("sp_sz", ""),
        }
        for p in range(1, 7):
            r[f"pos{p}_dim"]  = genes.get(f"pos{p}_dim",  "")
            r[f"pos{p}_tile"] = genes.get(f"pos{p}_tile", "")
        parsed_rows.append(r)

    headers2 = (
        ["Rank", "Freq", "Best Fitness", "Verdict", "SP Dim", "SP Size"] +
        [f"Loop{p}\nDim" for p in range(1, 7)] +
        [f"Loop{p}\nTile" for p in range(1, 7)]
    )
    write_header(ws2, headers2)

    col_keys = (["rank","frequency","best_fitness","verdict","sp_dim","sp_sz"] +
                [f"pos{p}_dim"  for p in range(1, 7)] +
                [f"pos{p}_tile" for p in range(1, 7)])

    for ri, r in enumerate(parsed_rows):
        er = ri + 2
        for ci, key in enumerate(col_keys, 1):
            c = ws2.cell(row=er, column=ci, value=r.get(key, ""))
            style_cell(c, r["verdict"])

    for i in range(1, len(headers2)+1):
        ws2.column_dimensions[get_column_letter(i)].width = 9
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 13
    ws2.column_dimensions["E"].width = 9
    ws2.column_dimensions["F"].width = 9
    ws2.freeze_panes = "A2"

    # ── Sheet 3: AVOID LIST ───────────────────────────────────────────────
    ws3 = wb.create_sheet("AVOID LIST")
    avoid = summary[summary["verdict"] == "BAD - AVOID"].copy()

    title_cell = ws3.cell(row=1, column=1,
        value="GENOMES TO AVOID IN FUTURE EXPLORATION — bottom 30% fitness (single layer)")
    title_cell.font = Font(bold=True, color="843C0C", name="Arial", size=11)
    title_cell.fill = PatternFill("solid", start_color="FCE4D6")
    title_cell.alignment = Alignment(horizontal="center")
    ws3.merge_cells("A1:F1")

    write_header(ws3, headers1, row=2)
    for ri, row in avoid.reset_index(drop=True).iterrows():
        er = ri + 3
        vals = [row["rank"], row["sp_dim_pattern"], row["frequency"],
                round(row["best_fitness"], 2) if row["best_fitness"] > -1e17 else "INVALID",
                row["verdict"], row["genome_str"]]
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(row=er, column=ci, value=val)
            style_cell(c, row["verdict"], wrap=(ci==6))

    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 14
    ws3.column_dimensions["F"].width = 80
    ws3.freeze_panes = "A3"

    wb.save(out_path)
    print("Saved -> {}".format(out_path))
    print("  Sheet 1 - Genome Summary : {} genomes".format(len(summary)))
    print("  Sheet 2 - Gene Detail    : {} rows (one col per gene position)".format(len(summary)))
    print("  Sheet 3 - Avoid List     : {} genomes to skip".format(len(avoid)))


def get_column_letter(n):
    result = ""
    while n:
        n, r = divmod(n-1, 26)
        result = chr(65+r) + result
    return result


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv",   required=True,               help="genome_all.csv from GAMMA run")
    parser.add_argument("--layer", type=int, default=1,         help="Which layer to analyze (default: 1)")
    parser.add_argument("--out",   default="genome_results.xlsx")
    args = parser.parse_args()

    print("Loading {} ...".format(args.csv))
    df = pd.read_csv(args.csv)

    if "layer" in df.columns:
        df = df[df["layer"] == args.layer].copy()
        print("Filtered to layer {}: {} rows".format(args.layer, len(df)))
    else:
        print("No layer column — using all {} rows".format(len(df)))

    summary = build_summary(df)
    print("Unique genome structures: {}".format(len(summary)))
    print("Total evaluations       : {}".format(len(df)))

    to_excel(summary, args.out)


if __name__ == "__main__":
    main()
